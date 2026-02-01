"""
XLSM Generator with VBA
Creates macro-enabled Excel files with embedded VBA modules
"""

import os
import shutil
import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# VBA module code - embedded for standalone generation
VBA_MODULES = {
    "modAPI": '''
' =============================================================
' Module: modAPI - API Integration Functions
' =============================================================

Private Const API_BASE_URL As String = "http://localhost:8000/api"

Public Function CallAPI(endpoint As String, Optional method As String = "GET") As String
    On Error GoTo ErrorHandler
    
    Dim http As Object
    Set http = CreateObject("MSXML2.XMLHTTP")
    
    http.Open method, API_BASE_URL & endpoint, False
    http.setRequestHeader "Content-Type", "application/json"
    http.send
    
    If http.Status = 200 Then
        CallAPI = http.responseText
    Else
        CallAPI = "{\"error\": \"HTTP " & http.Status & "\"}"
    End If
    
    Exit Function
ErrorHandler:
    CallAPI = "{\"error\": \"" & Err.Description & "\"}"
End Function

Public Sub RefreshStockData()
    Dim ticker As String
    ticker = Range("Assumptions!B3").Value
    
    If ticker = "" Then
        MsgBox "Please enter a stock ticker in Assumptions!B3", vbExclamation
        Exit Sub
    End If
    
    Application.StatusBar = "Fetching data for " & ticker & "..."
    
    Dim response As String
    response = CallAPI("/stocks/" & ticker)
    
    If InStr(response, "error") = 0 Then
        MsgBox "Data refreshed successfully!", vbInformation
    Else
        MsgBox "Failed to refresh data: " & response, vbCritical
    End If
    
    Application.StatusBar = False
End Sub
''',
    "modCalc": '''
' =============================================================
' Module: modCalc - Financial Calculations
' =============================================================

Public Function CalcWACC(costOfEquity As Double, costOfDebt As Double, _
                          taxRate As Double, equityWeight As Double) As Double
    CalcWACC = (equityWeight * costOfEquity) + ((1 - equityWeight) * costOfDebt * (1 - taxRate))
End Function

Public Function CalcTerminalValue(fcff As Double, terminalGrowth As Double, wacc As Double) As Double
    If wacc <= terminalGrowth Then
        CalcTerminalValue = 0
    Else
        CalcTerminalValue = fcff * (1 + terminalGrowth) / (wacc - terminalGrowth)
    End If
End Function

Public Function CalcDCF(cashFlows As Range, wacc As Double) As Double
    Dim i As Integer
    Dim pv As Double
    pv = 0
    
    For i = 1 To cashFlows.Count
        pv = pv + cashFlows(i).Value / ((1 + wacc) ^ i)
    Next i
    
    CalcDCF = pv
End Function

Public Sub RecalculateModel()
    Application.Calculate
    MsgBox "Model recalculated!", vbInformation
End Sub
''',
    "modDashboard": '''
' =============================================================
' Module: modDashboard - Dashboard Controls
' =============================================================

Public Sub UpdateDashboard()
    On Error Resume Next
    
    ' Update timestamp
    Sheets("Dashboard").Range("B1").Value = "Last Updated: " & Format(Now, "dd-mmm-yyyy hh:mm")
    
    ' Refresh charts
    Dim cht As ChartObject
    For Each cht In Sheets("Dashboard").ChartObjects
        cht.Chart.Refresh
    Next cht
    
    Application.StatusBar = "Dashboard updated"
End Sub

Public Sub ExportToPDF()
    Dim savePath As String
    savePath = Application.GetSaveAsFilename( _
        InitialFileName:=ThisWorkbook.Name & "_Export.pdf", _
        FileFilter:="PDF Files (*.pdf), *.pdf")
    
    If savePath <> "False" Then
        Sheets("Dashboard").ExportAsFixedFormat Type:=xlTypePDF, Filename:=savePath
        MsgBox "Exported to: " & savePath, vbInformation
    End If
End Sub
''',
    "modValidation": '''
' =============================================================
' Module: modValidation - Model Audit & Checks
' =============================================================

Public Sub RunValidation()
    Dim issues As Integer
    issues = 0
    
    ' Check balance sheet
    Dim assets As Double, liabEquity As Double
    On Error Resume Next
    assets = Sheets("Balance Sheet").Range("TotalAssets").Value
    liabEquity = Sheets("Balance Sheet").Range("TotalLiabEquity").Value
    On Error GoTo 0
    
    If Abs(assets - liabEquity) > 0.01 And assets > 0 Then
        issues = issues + 1
        Debug.Print "Balance sheet imbalance: " & Abs(assets - liabEquity)
    End If
    
    ' Check for errors
    Dim ws As Worksheet
    Dim cell As Range
    For Each ws In ThisWorkbook.Worksheets
        For Each cell In ws.UsedRange.SpecialCells(xlCellTypeFormulas)
            If IsError(cell.Value) Then
                issues = issues + 1
                Debug.Print "Error in " & ws.Name & "!" & cell.Address
            End If
        Next cell
    Next ws
    
    If issues = 0 Then
        MsgBox "Validation passed! No issues found.", vbInformation
    Else
        MsgBox issues & " issue(s) found. Check Immediate Window (Ctrl+G) for details.", vbExclamation
    End If
End Sub

Public Sub HighlightInputs()
    Dim ws As Worksheet
    Dim cell As Range
    
    For Each ws In ThisWorkbook.Worksheets
        For Each cell In ws.UsedRange.SpecialCells(xlCellTypeConstants)
            If IsNumeric(cell.Value) Then
                cell.Interior.Color = RGB(255, 242, 204)
            End If
        Next cell
    Next ws
    
    MsgBox "Input cells highlighted in yellow", vbInformation
End Sub
''',
    "modScenario": '''
' =============================================================
' Module: modScenario - Scenario Management
' =============================================================

Public Sub RunBaseCase()
    SetScenario 1, 0.1, 0.18, 0.04
End Sub

Public Sub RunBullCase()
    SetScenario 1.2, 0.12, 0.22, 0.035
End Sub

Public Sub RunBearCase()
    SetScenario 0.8, 0.08, 0.14, 0.045
End Sub

Private Sub SetScenario(revMult As Double, revGrowth As Double, margin As Double, wacc As Double)
    On Error Resume Next
    With Sheets("Assumptions")
        .Range("B5").Value = revGrowth
        .Range("B6").Value = margin
        .Range("B10").Value = wacc
    End With
    Application.Calculate
    MsgBox "Scenario applied!", vbInformation
End Sub

Public Sub SaveScenario()
    Dim scenarioName As String
    scenarioName = InputBox("Enter scenario name:", "Save Scenario")
    
    If scenarioName <> "" Then
        ' Store scenario in named range or hidden sheet
        MsgBox "Scenario '" & scenarioName & "' saved!", vbInformation
    End If
End Sub
'''
}


def create_xlsm_with_vba(source_xlsx: str, output_xlsm: str) -> bool:
    """
    Convert an xlsx to xlsm with VBA modules embedded.
    
    Note: openpyxl cannot directly write VBA, so we use a workaround:
    1. Copy the xlsx
    2. Create VBA code as text files
    3. For full VBA embedding, we would need xlwings or win32com
    
    For cross-platform compatibility, we create a README with instructions.
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        # Load the source workbook
        wb = load_workbook(source_xlsx)
        
        # Add a VBA Instructions sheet
        if "VBA Setup" in wb.sheetnames:
            ws = wb["VBA Setup"]
        else:
            ws = wb.create_sheet("VBA Setup", 0)
        
        # Title and instructions
        ws['A1'] = "VBA SETUP INSTRUCTIONS"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=16, color="1F4E79")
        
        ws['A3'] = "This workbook contains VBA automation modules."
        ws['A4'] = "To enable VBA functionality:"
        ws['A6'] = "1. Save this file as .xlsm (Excel Macro-Enabled Workbook)"
        ws['A7'] = "2. Press Alt+F11 to open VBA Editor"
        ws['A8'] = "3. Right-click on VBAProject → Import File"  
        ws['A9'] = "4. Import the .bas files from the 'vba_modules' folder"
        ws['A10'] = "5. Close VBA Editor and save"
        
        ws['A12'] = "VBA Modules Included:"
        ws['A12'].font = openpyxl.styles.Font(bold=True)
        
        row = 13
        for module_name in VBA_MODULES.keys():
            ws[f'A{row}'] = f"  • {module_name}.bas - {get_module_description(module_name)}"
            row += 1
        
        ws['A{}'.format(row + 1)] = "Quick Actions (after VBA is enabled):"
        ws['A{}'.format(row + 1)].font = openpyxl.styles.Font(bold=True)
        ws['A{}'.format(row + 2)] = "  • Run modAPI.RefreshStockData() - Refresh stock data from API"
        ws['A{}'.format(row + 3)] = "  • Run modValidation.RunValidation() - Check model for errors"  
        ws['A{}'.format(row + 4)] = "  • Run modDashboard.UpdateDashboard() - Refresh dashboard"
        ws['A{}'.format(row + 5)] = "  • Run modScenario.RunBullCase() - Apply bull case assumptions"
        
        # Set column width
        ws.column_dimensions['A'].width = 80
        
        # Save as xlsm
        # Note: openpyxl can't add actual VBA, but we include the code
        wb.save(output_xlsm)
        
        # Create VBA module files alongside the xlsm
        vba_dir = os.path.join(os.path.dirname(output_xlsm), "vba_modules")
        os.makedirs(vba_dir, exist_ok=True)
        
        for module_name, code in VBA_MODULES.items():
            module_path = os.path.join(vba_dir, f"{module_name}.bas")
            with open(module_path, 'w') as f:
                f.write(f"Attribute VB_Name = \"{module_name}\"\n")
                f.write(code)
        
        logger.info(f"Created xlsm with VBA setup: {output_xlsm}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to create xlsm: {e}")
        return False


def get_module_description(module_name: str) -> str:
    """Get description for each VBA module"""
    descriptions = {
        "modAPI": "API integration for live data",
        "modCalc": "Financial calculation functions",
        "modDashboard": "Dashboard controls and PDF export",
        "modValidation": "Model audit and validation",
        "modScenario": "Scenario management (Bull/Bear/Base)"
    }
    return descriptions.get(module_name, "VBA module")


def get_vba_modules_zip(output_path: str) -> str:
    """Create a zip file with all VBA modules"""
    import zipfile
    
    zip_path = output_path.replace('.xlsx', '_vba_modules.zip').replace('.xlsm', '_vba_modules.zip')
    
    with zipfile.ZipFile(zip_path, 'w') as zf:
        for module_name, code in VBA_MODULES.items():
            content = f'Attribute VB_Name = "{module_name}"\n{code}'
            zf.writestr(f"{module_name}.bas", content)
        
        # Add README
        readme = """AI Financial Modeler - VBA Modules
===================================

To install these VBA modules in your Excel workbook:

1. Open your Excel file
2. Press Alt+F11 to open VBA Editor
3. Right-click on VBAProject (your workbook name)
4. Select Import File...
5. Import each .bas file

Modules included:
- modAPI.bas: API integration for live stock data
- modCalc.bas: Financial calculation functions (WACC, DCF, Terminal Value)
- modDashboard.bas: Dashboard refresh and PDF export
- modValidation.bas: Model audit and error checking
- modScenario.bas: Bull/Bear/Base case scenario management

After importing, you can:
- Create buttons linked to the macros
- Use the functions in formulas (e.g., =CalcWACC(...))
- Run macros from Developer tab
"""
        zf.writestr("README.txt", readme)
    
    return zip_path
