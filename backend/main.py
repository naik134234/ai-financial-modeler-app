"""
AI Financial Modeling Platform - Main API Server
FastAPI backend for generating institutional-grade Excel financial models
"""

from fastapi import FastAPI, HTTPException, BackgroundTasks, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
import os
import logging
from datetime import datetime
import uuid
import tempfile
import openpyxl
import asyncio

import os
import sys

# Fix for Vercel: Add current directory to sys.path so imports work
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Import our modules
from data import fetch_stock_data, fetch_screener_data
from data.stock_database import get_all_stocks, get_stocks_by_sector, search_stocks, get_sectors
from agents import classify_company, create_model_structure, validate_financial_model
from excel import generate_financial_model
from excel.lbo_generator import generate_lbo_model
from excel.ma_generator import generate_ma_model

# Import new enhanced modules
import database as db
from agents.ai_assistant import generate_smart_assumptions, generate_valuation_commentary, parse_natural_language_request
from data.yahoo_finance import get_stock_info, get_historical_financials, get_price_history
from exporters import pdf_exporter, pptx_exporter
from analysis.monte_carlo import run_monte_carlo_simulation
from data.damodaran_data import get_all_industry_data, map_yahoo_industry, get_india_erp
from data.alpha_vantage import fetch_alpha_vantage_data, AlphaVantageAPI

# Import Chat and Analysis modules
from agents.chat_assistant import process_chat_message
from analysis.tornado_analysis import calculate_sensitivity
from analysis.football_field import create_football_field

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
logger.info("--- SERVER STARTED WITH CHAT & CHARTS ENABLED ---")

# Initialize FastAPI app
app = FastAPI(
    title="AI Financial Modeling Platform",
    description="Generate institutional-grade Excel financial models with AI",
    version="1.0.0",
)

# CORS middleware for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000", 
        "http://127.0.0.1:3000", 
        "http://localhost:3001", 
        "http://127.0.0.1:3001",
        "https://ai-financial-modeler.vercel.app",
        "https://ai-financial-modeler-backend-production.up.railway.app"
    ],
    allow_origin_regex=r"https://ai-financial-modeler.*\.vercel\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Output directory for generated models
# Use /tmp on Vercel (read-only filesystem elsewhere)
if os.environ.get("VERCEL"):
    OUTPUT_DIR = os.path.join(tempfile.gettempdir(), "output")
else:
    OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")

os.makedirs(OUTPUT_DIR, exist_ok=True)


# Pydantic models for API
class ModelRequest(BaseModel):
    """Request to generate a financial model"""
    symbol: str = Field(..., description="Stock symbol (e.g., ADANIPOWER)")
    exchange: str = Field(default="NSE", description="Exchange: NSE or BSE")
    forecast_years: int = Field(default=5, ge=1, le=10, description="Number of forecast years")
    model_types: List[str] = Field(
        default=["three_statement", "dcf"],
        description="Model types to include"
    )

class ModelResponse(BaseModel):
    """Response with model generation status"""
    job_id: str
    status: str
    message: str
    company_name: Optional[str] = None
    industry: Optional[str] = None
    download_url: Optional[str] = None

class CompanyInfo(BaseModel):
    """Company information response"""
    symbol: str
    name: str
    sector: str
    industry: str
    market_cap: Optional[float] = None
    current_price: Optional[float] = None

class ValidationResult(BaseModel):
    """Model validation result"""
    is_valid: bool
    errors: List[Dict[str, Any]]

class RawDataRequest(BaseModel):
    """Request to generate model from raw data input"""
    company_name: str = Field(..., description="Name of the company")
    industry: str = Field(default="general", description="Industry code")
    forecast_years: int = Field(default=5, ge=1, le=10)
    historical_data: Dict[str, Any] = Field(
        default={},
        description="Historical financial data (revenue, ebitda, net_income, etc.)"
    )
    assumptions: Dict[str, float] = Field(
        default={},
        description="Override default assumptions (revenue_growth, ebitda_margin, etc.)"
    )


class LBORequest(BaseModel):
    """Request to generate an LBO (Leveraged Buyout) model"""
    symbol: str = Field(..., description="Stock symbol (e.g., ADANIPOWER)")
    exchange: str = Field(default="NSE", description="Exchange: NSE or BSE")
    holding_period: int = Field(default=5, ge=3, le=10, description="Holding period in years")
    
    # Transaction assumptions
    entry_multiple: float = Field(default=8.0, ge=3.0, le=20.0, description="Entry EV/EBITDA multiple")
    exit_multiple: float = Field(default=8.0, ge=3.0, le=20.0, description="Exit EV/EBITDA multiple")
    transaction_fees_pct: float = Field(default=0.02, ge=0, le=0.10, description="Transaction fees as % of EV")
    financing_fees_pct: float = Field(default=0.03, ge=0, le=0.10, description="Financing fees as % of debt")
    management_rollover_pct: float = Field(default=0.10, ge=0, le=0.50, description="Management rollover %")
    
    # Debt structure
    senior_debt_multiple: float = Field(default=3.0, ge=0, le=6.0, description="Senior debt x EBITDA")
    senior_interest_rate: float = Field(default=0.08, ge=0.03, le=0.20, description="Senior debt interest rate")
    senior_amort_years: int = Field(default=7, ge=5, le=10, description="Senior debt amortization years")
    mezz_debt_multiple: float = Field(default=1.5, ge=0, le=3.0, description="Mezzanine debt x EBITDA")
    mezz_interest_rate: float = Field(default=0.12, ge=0.05, le=0.25, description="Mezzanine cash interest rate")
    mezz_pik_rate: float = Field(default=0.02, ge=0, le=0.10, description="Mezzanine PIK interest rate")
    sub_debt_multiple: float = Field(default=0.5, ge=0, le=2.0, description="Subordinated debt x EBITDA")
    sub_interest_rate: float = Field(default=0.14, ge=0.05, le=0.25, description="Subordinated debt interest rate")
    
    # Operating assumptions
    revenue_growth: float = Field(default=0.08, ge=-0.20, le=0.50, description="Annual revenue growth rate")
    ebitda_margin: float = Field(default=0.25, ge=0.05, le=0.60, description="EBITDA margin")
    capex_pct: float = Field(default=0.04, ge=0, le=0.20, description="CapEx as % of revenue")
    nwc_pct: float = Field(default=0.10, ge=0, le=0.30, description="Net working capital as % of revenue")
    tax_rate: float = Field(default=0.25, ge=0.10, le=0.40, description="Corporate tax rate")


class MARequest(BaseModel):
    """Request to generate an M&A (Merger & Acquisition) model"""
    # Company symbols
    acquirer_symbol: str = Field(..., description="Acquirer stock symbol")
    target_symbol: str = Field(..., description="Target stock symbol")
    exchange: str = Field(default="NSE", description="Exchange: NSE or BSE")
    
    # Transaction assumptions
    offer_premium: float = Field(default=0.25, ge=0, le=1.0, description="Offer premium over target price")
    percent_stock: float = Field(default=0.50, ge=0, le=1.0, description="% stock consideration")
    percent_cash: float = Field(default=0.50, ge=0, le=1.0, description="% cash consideration")
    transaction_fees_pct: float = Field(default=0.02, ge=0, le=0.10, description="Transaction fees %")
    financing_rate: float = Field(default=0.06, ge=0.03, le=0.15, description="Debt financing rate")
    
    # Synergy assumptions
    synergies_revenue: float = Field(default=0, ge=0, description="Annual revenue synergies")
    synergies_cost: float = Field(default=0, ge=0, description="Annual cost synergies")
    integration_costs: float = Field(default=0, ge=0, description="One-time integration costs")
    synergy_phase_in_year1: float = Field(default=0.25, ge=0, le=1, description="Synergy realization Year 1")
    synergy_phase_in_year2: float = Field(default=0.50, ge=0, le=1, description="Synergy realization Year 2")
    synergy_phase_in_year3: float = Field(default=1.00, ge=0, le=1, description="Synergy realization Year 3")
    
    # Growth assumptions
    acquirer_growth_rate: float = Field(default=0.05, ge=-0.2, le=0.5, description="Acquirer revenue growth")
    target_growth_rate: float = Field(default=0.05, ge=-0.2, le=0.5, description="Target revenue growth")


# Job storage with database persistence
from job_manager import jobs


@app.get("/")
async def root():
    """Root endpoint with API info"""
    return {
        "name": "AI Financial Modeling Platform",
        "version": "1.0.0",
        "status": "running",
        "endpoints": {
            "company_info": "/api/company/{symbol}",
            "generate_model": "/api/model/generate",
            "job_status": "/api/job/{job_id}",
            "download_model": "/api/download/{job_id}",
        }
    }


@app.get("/api/company/{symbol}", response_model=CompanyInfo)
async def get_company_info(symbol: str, exchange: str = "NSE"):
    """
    Get company information for a stock symbol
    
    Args:
        symbol: Stock symbol (e.g., ADANIPOWER, RELIANCE)
        exchange: NSE or BSE
    """
    try:
        # Fetch data from Yahoo Finance
        data = await fetch_stock_data(symbol, exchange)
        info = data.get('company_info', {})
        metrics = data.get('key_metrics', {})
        
        return CompanyInfo(
            symbol=symbol.upper(),
            name=info.get('name', symbol),
            sector=info.get('sector', 'Unknown'),
            industry=info.get('industry', 'Unknown'),
            market_cap=info.get('market_cap'),
            current_price=metrics.get('current_price'),
        )
    except Exception as e:
        logger.error(f"Error fetching company info for {symbol}: {e}")
        raise HTTPException(status_code=404, detail=f"Company not found: {symbol}")

        raise HTTPException(status_code=404, detail=f"Company not found: {symbol}")


# --- NEW API ROUTES FOR CHAT & ANALYSIS ---

class ChatRequest(BaseModel):
    message: str
    history: Optional[List[Dict[str, str]]] = None

@app.post("/api/chat/{job_id}")
async def chat_endpoint(job_id: str, request: ChatRequest):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    
    # Ensure job has data needed for chat
    # We pass the whole job dict, the assistant handles extraction
    
    # Run in threadpool to avoid blocking
    loop = asyncio.get_event_loop()
    response = await loop.run_in_executor(None, process_chat_message, request.message, job, request.history)
    
    if not response['success']:
        # Return 200 with error message as chat response so UI doesn't crash
        return {"response": response.get('response', 'Error processing request')}
        
    return {"response": response['response']}

@app.get("/api/analysis/sensitivity/{job_id}")
async def get_sensitivity(job_id: str):
    if job_id not in jobs:
         raise HTTPException(status_code=404, detail="Job not found")
    job = jobs[job_id]
    
    # Check if we have valuation data
    # (assuming job stores 'valuation_summary' or similar from generate_financial_model)
    # If not present, we might need to fallback or check job status
    
    valuation_data = job.get('valuation_data', {})
    assumptions = job.get('assumptions', {})
    
    if not valuation_data:
        # Fallback: try to construct minimal data from request if job finished?
        # Or return empty
        pass
        
    data = calculate_sensitivity(valuation_data, assumptions)
    return {"sensitivity": data}

@app.get("/api/analysis/football-field/{job_id}")
async def get_football_field(job_id: str):
    if job_id not in jobs:
         raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    
    valuation_data = job.get('valuation_data', {})
    monte_carlo = job.get('monte_carlo_results')
    competitors = job.get('competitors_data', {}) # or similar
    
    data = create_football_field(valuation_data, monte_carlo, competitors)
    return {"football_field": data}


@app.post("/api/model/generate", response_model=ModelResponse)
async def generate_model(request: ModelRequest, background_tasks: BackgroundTasks):
    """
    Generate a financial model for a company
    
    This starts an async job and returns a job ID for tracking.
    """
    job_id = str(uuid.uuid4())
    
    # Initialize job
    jobs[job_id] = {
        "status": "pending",
        "created_at": datetime.now().isoformat(),
        "request": request.dict(),
        "progress": 0,
        "message": "Job queued",
    }
    
    # Start background task
    if os.environ.get("VERCEL"):
        # On Vercel, run synchronously to prevent timeout/kill
        await _generate_model_task(
            job_id,
            request.symbol,
            request.exchange,
            request.forecast_years,
        )
    else:
        background_tasks.add_task(
            _generate_model_task,
            job_id,
            request.symbol,
            request.exchange,
            request.forecast_years,
        )
    
    return ModelResponse(
        job_id=job_id,
        status="pending",
        message="Model generation started. Check job status for progress.",
    )


def _extract_screener_financials(annual_results: List[Dict]) -> Dict[str, Any]:
    """
    Extract real financial figures from Screener.in annual results
    
    Args:
        annual_results: List of dicts with 'metric' and year columns
    
    Returns:
        Dictionary with extracted financials (revenue, ebitda, net_income, etc.)
    """
    financials = {}
    
    # Build a lookup by metric name
    metric_map = {}
    for row in annual_results:
        metric = row.get('metric', '').lower().strip()
        if metric:
            metric_map[metric] = row
    
    # Common metric name mappings
    mappings = {
        'revenue': ['sales', 'revenue', 'total revenue', 'net sales', 'income from operations'],
        'expenses': ['expenses', 'total expenses', 'operating expenses'],
        'operating_profit': ['operating profit', 'ebit', 'operating income'],
        'ebitda': ['ebitda', 'operating profit before interest'],
        'interest': ['interest', 'finance costs', 'interest expense'],
        'net_income': ['net profit', 'net income', 'profit after tax', 'pat'],
        'eps': ['eps', 'earnings per share', 'basic eps'],
        'dividend': ['dividend payout', 'dividend', 'dividend %'],
    }
    
    # Get the last year column (most recent)
    year_columns = []
    if annual_results:
        first_row = annual_results[0]
        year_columns = [k for k in first_row.keys() if k not in ['metric'] and 'mar' in k.lower() or k.isdigit()]
        if not year_columns:
            # Try to find any numeric-looking columns
            year_columns = [k for k in first_row.keys() if k not in ['metric']]
    
    latest_year = year_columns[-1] if year_columns else None
    
    for fin_key, possible_names in mappings.items():
        for name in possible_names:
            if name in metric_map:
                row = metric_map[name]
                # Get most recent year value
                if latest_year and latest_year in row:
                    financials[fin_key] = row[latest_year]
                # Also store historical values
                financials[f'{fin_key}_history'] = {k: v for k, v in row.items() if k != 'metric'}
                break
    
    # Calculate margins from the data
    if financials.get('revenue') and financials.get('operating_profit'):
        try:
            rev = float(financials['revenue']) if financials['revenue'] else 0
            op = float(financials['operating_profit']) if financials['operating_profit'] else 0
            if rev > 0:
                financials['operating_margin'] = op / rev
        except (ValueError, TypeError):
            pass
    
    if financials.get('revenue') and financials.get('net_income'):
        try:
            rev = float(financials['revenue']) if financials['revenue'] else 0
            ni = float(financials['net_income']) if financials['net_income'] else 0
            if rev > 0:
                financials['net_margin'] = ni / rev
        except (ValueError, TypeError):
            pass
    
    return financials


async def _generate_model_task(
    job_id: str,
    symbol: str,
    exchange: str,
    forecast_years: int,
):
    """Background task to generate the financial model"""
    try:
        jobs[job_id]["status"] = "processing"
        jobs[job_id]["progress"] = 5
        jobs[job_id]["message"] = "Fetching financial data from multiple sources in parallel..."
        
        # Helper wrappers for sync functions
        loop = asyncio.get_event_loop()
        
        async def fetch_av():
            try:
                logger.info(f"Fetching Alpha Vantage data for {symbol}")
                return await loop.run_in_executor(None, fetch_alpha_vantage_data, symbol)
            except Exception as e:
                logger.warning(f"Alpha Vantage fetch failed: {e}")
                return {}

        async def fetch_sc():
            try:
                return await loop.run_in_executor(None, fetch_screener_data, symbol)
            except Exception as e:
                logger.warning(f"Screener scraping failed: {e}")
                return {}

        async def fetch_yf():
            try:
                logger.info(f"Fetching Yahoo Finance data for {symbol}")
                return await fetch_stock_data(symbol, exchange)
            except Exception as e:
                logger.warning(f"Yahoo fetch failed: {e}")
                return {}

        # 1. Start parallel fetching for Company Data
        av_task = asyncio.create_task(fetch_av())
        sc_task = asyncio.create_task(fetch_sc())
        yf_task = asyncio.create_task(fetch_yf())
        
        alpha_vantage_data, screener_data, yahoo_data = await asyncio.gather(av_task, sc_task, yf_task)
        
        # Log success
        if alpha_vantage_data.get('company_info'):
            logger.info("Alpha Vantage: Got real data")
        if screener_data.get('annual_results'):
            logger.info(f"Screener data fetched: {len(screener_data['annual_results'])} records")
        
        jobs[job_id]["progress"] = 35
        jobs[job_id]["message"] = "Fetching Damodaran industry benchmarks..."
        
        # 2. Fetch Damodaran data (depends on Yahoo industry)
        company_info = yahoo_data.get('company_info', {})
        yahoo_industry = company_info.get('industry', 'Unknown')
        damodaran_industry = map_yahoo_industry(yahoo_industry)
        
        try:
            damodaran_data = await loop.run_in_executor(None, get_all_industry_data, damodaran_industry)
            logger.info(f"Damodaran data fetched for: {damodaran_industry}")
        except Exception as e:
            logger.warning(f"Damodaran fetch failed: {e}")
            damodaran_data = {}
        
        # Step 4: Merge all data sources with priority
        # Priority: Alpha Vantage > Screener.in > Yahoo Finance > Damodaran defaults
        financial_data = {**yahoo_data}
        
        # HIGHEST PRIORITY: Alpha Vantage real data (has accurate shares, market cap, fundamentals)
        if alpha_vantage_data:
            av_company_info = alpha_vantage_data.get('company_info', {})
            if av_company_info:
                # Merge Alpha Vantage data into company_info
                yahoo_company_info = financial_data.get('company_info', {})
                merged_company_info = {**yahoo_company_info, **av_company_info}
                financial_data['company_info'] = merged_company_info
                company_info = merged_company_info
                logger.info(f"Merged Alpha Vantage data - Shares: {av_company_info.get('shares_outstanding')}, Market Cap: {av_company_info.get('market_cap')}")
            
            # Add Alpha Vantage financial statements
            if alpha_vantage_data.get('income_statement'):
                financial_data['av_income_statement'] = alpha_vantage_data.get('income_statement')
            if alpha_vantage_data.get('balance_sheet'):
                financial_data['av_balance_sheet'] = alpha_vantage_data.get('balance_sheet')
            if alpha_vantage_data.get('cash_flow'):
                financial_data['av_cash_flow'] = alpha_vantage_data.get('cash_flow')
            
            # Use Alpha Vantage real financials if available
            if alpha_vantage_data.get('real_financials'):
                financial_data['real_financials'] = alpha_vantage_data['real_financials']
                logger.info(f"Using Alpha Vantage real financials: {list(alpha_vantage_data['real_financials'].keys())}")
        
        # Add Screener.in data (supplements Alpha Vantage)
        if screener_data:
            # Merge Screener.in company info (if Alpha Vantage didn't provide some fields)
            screener_company_info = screener_data.get('company_info', {})
            if screener_company_info:
                # Only add fields not already present from Alpha Vantage
                current_company_info = financial_data.get('company_info', {})
                for key, value in screener_company_info.items():
                    if key not in current_company_info or not current_company_info[key]:
                        current_company_info[key] = value
                financial_data['company_info'] = current_company_info
                company_info = current_company_info
                logger.info(f"Supplemented with Screener.in company info")
            
            financial_data['screener_annual'] = screener_data.get('annual_results', [])
            financial_data['screener_quarterly'] = screener_data.get('quarterly_results', [])
            financial_data['screener_balance_sheet'] = screener_data.get('balance_sheet', [])
            financial_data['screener_cash_flow'] = screener_data.get('cash_flow', [])
            financial_data['screener_ratios'] = screener_data.get('ratios', [])
            financial_data['peers'] = screener_data.get('peers', [])
            
            # Extract real historical financials from Screener if not already from Alpha Vantage
            if not financial_data.get('real_financials'):
                annual = screener_data.get('annual_results', [])
                if annual:
                    real_financials = _extract_screener_financials(annual)
                    financial_data['real_financials'] = real_financials
                    logger.info(f"Using Screener.in financials: Revenue={real_financials.get('revenue')}, Net Income={real_financials.get('net_income')}")
        
        # Add Damodaran assumptions for projections (fallback for industry benchmarks)
        if damodaran_data:
            financial_data['damodaran'] = damodaran_data
            financial_data['model_assumptions'] = damodaran_data.get('model_assumptions', {})
            financial_data['data_source'] = f"Primary: Alpha Vantage API | Secondary: Screener.in + Yahoo | Industry: Damodaran ({damodaran_industry})"
        
        jobs[job_id]["progress"] = 45
        jobs[job_id]["message"] = "Classifying industry..."
        
        # Step 5: Classify industry
        industry_info = classify_company(company_info)
        
        # Enhance industry_info with Damodaran data
        if damodaran_data:
            industry_info['damodaran_industry'] = damodaran_industry
            industry_info['industry_beta'] = damodaran_data.get('beta', {}).get('levered_beta', 1.0)
            industry_info['industry_wacc'] = damodaran_data.get('wacc', {}).get('wacc', 0.11)
        
        jobs[job_id]["company_name"] = company_info.get('name', symbol)
        jobs[job_id]["industry"] = industry_info.get('industry_name', 'Unknown')
        jobs[job_id]["progress"] = 55
        jobs[job_id]["message"] = "Designing model structure..."
        
        # Step 6: Design model structure with real assumptions
        model_structure = create_model_structure(
            company_name=company_info.get('name', symbol),
            industry_info=industry_info,
            historical_data=financial_data,
            forecast_years=forecast_years,
        )
        
        # Override assumptions with Damodaran data
        if damodaran_data and 'model_assumptions' in damodaran_data:
            model_structure['assumptions'] = {
                **model_structure.get('assumptions', {}),
                **damodaran_data['model_assumptions']
            }
        
        jobs[job_id]["progress"] = 70
        jobs[job_id]["message"] = "Generating Excel model with real data..."
        
        # Step 7: Generate Excel file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{symbol}_{industry_info['model_type']}_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        
        generate_financial_model(
            company_name=company_info.get('name', symbol),
            model_structure=model_structure,
            financial_data=financial_data,
            industry_info=industry_info,
            output_path=output_path,
        )
        
        jobs[job_id]["progress"] = 90
        jobs[job_id]["message"] = "Validating model..."
        
        # Step 5: Validate the model
        validation_data = {
            'income_statement': financial_data.get('income_statement', {}),
            'balance_sheet': financial_data.get('balance_sheet', {}),
            'cash_flow': financial_data.get('cash_flow', {}),
            'assumptions': {},
        }
        is_valid, errors = validate_financial_model(
            validation_data,
            industry_info.get('industry_code', 'general')
        )
        
        # Complete
        jobs[job_id]["status"] = "completed"
        jobs[job_id]["progress"] = 100
        jobs[job_id]["message"] = "Model generated successfully!"
        jobs[job_id]["file_path"] = output_path
        jobs[job_id]["filename"] = filename
        jobs[job_id]["validation"] = {
            "is_valid": is_valid,
            "errors": errors[:5] if errors else [],  # Limit to first 5 errors
        }
        jobs[job_id]["download_url"] = f"/api/download/{job_id}"
        
        logger.info(f"Model generated successfully: {filename}")
        
    except Exception as e:
        logger.error(f"Error generating model for job {job_id}: {e}")
        jobs[job_id]["status"] = "failed"
        jobs[job_id]["message"] = f"Error: {str(e)}"
        jobs[job_id]["progress"] = 0


@app.get("/api/job/{job_id}")
async def get_job_status(job_id: str):
    """Get the status of a model generation job"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    
    response = {
        "job_id": job_id,
        "status": job["status"],
        "progress": job["progress"],
        "message": job["message"],
        "company_name": job.get("company_name"),
        "industry": job.get("industry"),
    }
    
    if job["status"] == "completed":
        response["download_url"] = job.get("download_url")
        response["filename"] = job.get("filename")
        response["validation"] = job.get("validation")
    
    return response


@app.get("/api/download/{job_id}")
async def download_model(job_id: str):
    """Download the generated Excel model"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    
    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail="Model not ready for download")
    
    file_path = job.get("file_path")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=file_path,
        filename=job.get("filename", "financial_model.xlsx"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/industries")
async def get_industries():
    """Get list of supported industries"""
    from agents.industry_classifier import INDUSTRY_TEMPLATES
    
    industries = []
    for code, template in INDUSTRY_TEMPLATES.items():
        industries.append({
            "code": code,
            "name": template["name"],
            "model_type": template["model_type"],
            "key_metrics": template["key_metrics"],
        })
    
    return {"industries": industries}


# ======================= DAMODARAN DATA API =======================

@app.get("/api/damodaran/industries")
async def get_damodaran_industries():
    """Get list of industries from Damodaran's database"""
    from data.damodaran_data import list_available_industries
    
    try:
        industries = list_available_industries()
        return {
            "source": "Damodaran Online (pages.stern.nyu.edu/~adamodar/)",
            "industries": industries,
            "count": len(industries)
        }
    except Exception as e:
        logger.error(f"Failed to fetch Damodaran industries: {e}")
        return {"industries": [], "error": str(e)}


@app.get("/api/damodaran/assumptions/{industry}")
async def get_damodaran_assumptions(industry: str):
    """
    Get Damodaran industry assumptions for valuation
    
    Args:
        industry: Industry name (e.g., 'Oil/Gas', 'Banking', 'Computer Software')
    
    Returns:
        Complete set of industry assumptions including beta, WACC, margins, multiples
    """
    try:
        data = get_all_industry_data(industry)
        return data
    except Exception as e:
        logger.error(f"Failed to fetch Damodaran data for {industry}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch industry data: {str(e)}")


@app.get("/api/damodaran/erp")
async def get_equity_risk_premium():
    """Get India-specific equity risk premium from Damodaran"""
    try:
        erp_data = get_india_erp()
        return {
            "source": "Damodaran Online",
            "country": "India",
            **erp_data
        }
    except Exception as e:
        logger.error(f"Failed to fetch ERP data: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ======================= COMPARISON API =======================

class CompareRequest(BaseModel):
    symbols: List[str] = Field(..., min_items=2, max_items=5, description="List of stock symbols to compare")
    exchange: str = Field(default="NSE", description="Stock exchange")


@app.post("/api/compare")
async def compare_companies(request: CompareRequest):
    """Compare multiple companies side-by-side"""
    try:
        comparison_data = []
        
        for symbol in request.symbols:
            try:
                # Fetch data for each company
                stock_data = await fetch_stock_data(symbol, request.exchange)
                screener_data = await fetch_screener_data(symbol)
                
                company_info = stock_data.get('company_info', {})
                financials = stock_data.get('financials', {})
                
                # Extract key metrics for comparison
                company_metrics = {
                    "symbol": symbol,
                    "name": company_info.get('name', symbol),
                    "sector": company_info.get('sector', 'Unknown'),
                    "market_cap": company_info.get('market_cap', 0),
                    "current_price": company_info.get('current_price', 0),
                    
                    # Valuation metrics
                    "pe_ratio": financials.get('pe_ratio', 0),
                    "pb_ratio": financials.get('pb_ratio', 0),
                    "ev_ebitda": financials.get('ev_ebitda', 0),
                    
                    # Profitability
                    "revenue": financials.get('revenue', 0),
                    "ebitda": financials.get('ebitda', 0),
                    "net_income": financials.get('net_income', 0),
                    "gross_margin": financials.get('gross_margin', 0),
                    "ebitda_margin": financials.get('ebitda_margin', 0),
                    "net_margin": financials.get('net_margin', 0),
                    "roe": financials.get('roe', 0),
                    "roce": financials.get('roce', 0),
                    
                    # Growth
                    "revenue_growth": financials.get('revenue_growth', 0),
                    "profit_growth": financials.get('profit_growth', 0),
                    
                    # Debt
                    "debt_to_equity": financials.get('debt_to_equity', 0),
                    "interest_coverage": financials.get('interest_coverage', 0),
                }
                
                comparison_data.append(company_metrics)
                
            except Exception as e:
                logger.warning(f"Failed to fetch data for {symbol}: {e}")
                comparison_data.append({
                    "symbol": symbol,
                    "name": symbol,
                    "error": str(e)
                })
        
        # Calculate averages for benchmarking
        metrics_to_average = ['pe_ratio', 'pb_ratio', 'ev_ebitda', 'gross_margin', 
                              'ebitda_margin', 'net_margin', 'roe', 'roce',
                              'revenue_growth', 'profit_growth', 'debt_to_equity']
        
        averages = {}
        for metric in metrics_to_average:
            values = [c.get(metric, 0) for c in comparison_data if c.get(metric) and not c.get('error')]
            if values:
                averages[metric] = sum(values) / len(values)
        
        return {
            "companies": comparison_data,
            "averages": averages,
            "count": len(comparison_data)
        }
        
    except Exception as e:
        logger.error(f"Comparison failed: {e}")
        raise HTTPException(status_code=500, detail=f"Comparison failed: {str(e)}")


@app.get("/api/compare/metrics")
async def get_comparison_metrics():
    """Get list of metrics available for comparison"""
    return {
        "metrics": [
            {"id": "pe_ratio", "name": "P/E Ratio", "format": "number", "category": "valuation"},
            {"id": "pb_ratio", "name": "P/B Ratio", "format": "number", "category": "valuation"},
            {"id": "ev_ebitda", "name": "EV/EBITDA", "format": "number", "category": "valuation"},
            {"id": "market_cap", "name": "Market Cap", "format": "currency", "category": "size"},
            {"id": "revenue", "name": "Revenue", "format": "currency", "category": "financials"},
            {"id": "ebitda", "name": "EBITDA", "format": "currency", "category": "financials"},
            {"id": "net_income", "name": "Net Income", "format": "currency", "category": "financials"},
            {"id": "gross_margin", "name": "Gross Margin", "format": "percent", "category": "profitability"},
            {"id": "ebitda_margin", "name": "EBITDA Margin", "format": "percent", "category": "profitability"},
            {"id": "net_margin", "name": "Net Margin", "format": "percent", "category": "profitability"},
            {"id": "roe", "name": "ROE", "format": "percent", "category": "profitability"},
            {"id": "roce", "name": "ROCE", "format": "percent", "category": "profitability"},
            {"id": "revenue_growth", "name": "Revenue Growth", "format": "percent", "category": "growth"},
            {"id": "profit_growth", "name": "Profit Growth", "format": "percent", "category": "growth"},
            {"id": "debt_to_equity", "name": "Debt/Equity", "format": "number", "category": "leverage"},
            {"id": "interest_coverage", "name": "Interest Coverage", "format": "number", "category": "leverage"},
        ]
    }


@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
    }


# ======================= TEMPLATES API =======================

@app.get("/api/templates/lbo")
async def get_lbo_templates():
    """Get all available LBO templates"""
    from templates import get_lbo_templates
    templates = get_lbo_templates()
    return {"templates": templates}


@app.get("/api/templates/lbo/{template_id}")
async def get_lbo_template(template_id: str):
    """Get a specific LBO template"""
    from templates import get_lbo_template
    template = get_lbo_template(template_id)
    if not template:
        raise HTTPException(status_code=404, detail=f"Template not found: {template_id}")
    return template


@app.get("/api/templates/ma")
async def get_ma_templates():
    """Get all available M&A templates"""
    from templates import get_ma_templates
    templates = get_ma_templates()
    return {"templates": templates}


@app.get("/api/templates/ma/{template_id}")
async def get_ma_template(template_id: str):
    """Get a specific M&A template"""
    from templates import get_ma_template
    template = get_ma_template(template_id)
    if not template:
        raise HTTPException(status_code=404, detail=f"Template not found: {template_id}")
    return template


# ======================= EXPORT API =======================

@app.get("/api/export/formats")
async def get_export_formats():
    """Get available export formats"""
    return {
        "formats": [
            {"id": "xlsx", "name": "Excel", "extension": ".xlsx", "available": True, "description": "Full financial model"},
            {"id": "pdf", "name": "PDF", "extension": ".pdf", "available": True, "description": "Executive summary report"},
            {"id": "pptx", "name": "PowerPoint", "extension": ".pptx", "available": True, "description": "Presentation slides"},
        ]
    }


@app.get("/api/export/{job_id}/pdf")
async def export_to_pdf(job_id: str):
    """Export model to PDF"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail="Model not ready for export")
    
    excel_path = job.get("file_path")
    if not excel_path or not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail="Model file not found")
    
    try:
        # Generate PDF using exporter
        pdf_path = excel_path.replace('.xlsx', '_report.pdf')
        pdf_exporter.create_pdf_report(
            company_name=job.get("company_name", "Company"),
            excel_path=excel_path,
            output_path=pdf_path
        )
        
        return FileResponse(
            path=pdf_path,
            filename=os.path.basename(pdf_path),
            media_type="application/pdf"
        )
    except Exception as e:
        logger.error(f"PDF export failed: {e}")
        raise HTTPException(status_code=500, detail=f"PDF export failed: {str(e)}")


@app.get("/api/export/{job_id}/pptx")
async def export_to_pptx(job_id: str):
    """Export model to PowerPoint"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail="Model not ready for export")
    
    excel_path = job.get("file_path")
    if not excel_path or not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail="Model file not found")
    
    try:
        # Generate PPTX using exporter
        pptx_path = excel_path.replace('.xlsx', '_presentation.pptx')
        pptx_exporter.create_presentation(
            company_name=job.get("company_name", "Company"),
            excel_path=excel_path,
            output_path=pptx_path
        )
        
        return FileResponse(
            path=pptx_path,
            filename=os.path.basename(pptx_path),
            media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation"
        )
    except Exception as e:
        logger.error(f"PPTX export failed: {e}")
        raise HTTPException(status_code=500, detail=f"PPTX export failed: {str(e)}")


# ======================= EXCEL PREVIEW API =======================

@app.post("/api/model/preview-excel")
async def preview_excel(file: UploadFile = File(...)):
    """Preview uploaded Excel file data before processing using openpyxl"""
    try:
        import io
        import openpyxl
        
        contents = await file.read()
        
        # Read Excel file using openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        
        preview_data = {
            "filename": file.filename,
            "sheets": wb.sheetnames,
            "data": {}
        }
        
        # Extract key financial data from sheets
        for sheet_name in wb.sheetnames:
            sheet_lower = sheet_name.lower()
            sheet = wb[sheet_name]
            
            # Helper to get first 5 rows and 5 columns
            rows = list(sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=5, values_only=True))
            if not rows:
                continue
                
            headers = [str(cell) if cell is not None else f"Col{i}" for i, cell in enumerate(rows[0])]
            data_rows = []
            for row in rows[1:]:
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = cell
                data_rows.append(row_dict)

            key = None
            if 'income' in sheet_lower or 'p&l' in sheet_lower or 'profit' in sheet_lower:
                key = "income_statement"
            elif 'balance' in sheet_lower:
                key = "balance_sheet"
            elif 'cash' in sheet_lower:
                key = "cash_flow"
            
            if key:
                preview_data["data"][key] = {
                    "rows": sheet.max_row,
                    "columns": headers,
                    "sample": data_rows
                }
        
        return preview_data
        
    except Exception as e:
        logger.error(f"Excel preview failed: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")


# ======================= MONTE CARLO API =======================

@app.get("/api/analysis/monte-carlo/{job_id}")
async def run_monte_carlo_endpoint(job_id: str, simulations: int = 1000):
    """Run Monte Carlo simulation on a completed model"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail="Model not ready for simulation")
    
    try:
        from analysis.monte_carlo import run_monte_carlo_simulation
        
        company_name = job.get("company_name", "Company")
        
        # Extract assumptions from job data or use defaults
        request_data = job.get("request", {})
        
        base_assumptions = {
            'revenue_growth': request_data.get('revenue_growth', 0.10),
            'ebitda_margin': request_data.get('ebitda_margin', 0.18),
            'terminal_growth': request_data.get('terminal_growth', 0.04),
            'wacc': request_data.get('wacc', 0.12),
        }
        
        # Extract valuation metrics from job or use estimates
        base_valuation = {
            'enterprise_value': job.get('valuation', {}).get('enterprise_value', 10000),
            'equity_value': job.get('valuation', {}).get('equity_value', 8000),
            'share_price': job.get('valuation', {}).get('share_price', 250),
            'net_debt': job.get('valuation', {}).get('net_debt', 2000),
        }
        
        # Run simulation (cap at 10k for performance)
        results = run_monte_carlo_simulation(
            base_assumptions=base_assumptions,
            base_valuation=base_valuation,
            num_simulations=min(simulations, 10000)
        )
        
        return {
            "job_id": job_id,
            "company_name": company_name,
            "simulations": min(simulations, 10000),
            "results": results
        }
    except Exception as e:
        logger.error(f"Monte Carlo simulation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Simulation failed: {str(e)}")



# ======================= CACHE API =======================

@app.get("/api/cache/stats")
async def get_cache_stats():
    """Get cache statistics"""
    try:
        from cache import get_cache_stats, clear_expired
        
        # Clear expired entries first
        cleared = clear_expired()
        
        stats = get_cache_stats()
        stats["cleared_expired"] = cleared
        
        return stats
    except Exception as e:
        return {"error": str(e), "total_entries": 0}


@app.delete("/api/cache/clear")
async def clear_cache():
    """Clear all cache entries"""
    try:
        from cache import clear_all
        cleared = clear_all()
        return {"cleared": cleared, "message": "Cache cleared successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ======================= SAVED PROJECTS API =======================

class SaveProjectRequest(BaseModel):
    name: str
    description: Optional[str] = None
    project_type: str = "general"  # 'general', 'lbo', 'ma'
    configuration: Dict[str, Any]


@app.get("/api/projects")
async def list_projects():
    """Get all saved projects"""
    from database import get_all_projects
    projects = get_all_projects()
    return {"projects": projects}


@app.get("/api/projects/{project_id}")
async def get_project_endpoint(project_id: int):
    """Get a specific project"""
    from database import get_project
    project = get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@app.post("/api/projects")
async def create_project(request: SaveProjectRequest):
    """Create a new saved project"""
    from database import save_project
    project = save_project(
        name=request.name,
        configuration=request.configuration,
        project_type=request.project_type,
        description=request.description
    )
    return project


@app.put("/api/projects/{project_id}")
async def update_project_endpoint(project_id: int, request: SaveProjectRequest):
    """Update an existing project"""
    from database import update_project
    project = update_project(
        project_id=project_id,
        name=request.name,
        configuration=request.configuration,
        description=request.description
    )
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@app.delete("/api/projects/{project_id}")
async def delete_project_endpoint(project_id: int):
    """Delete a saved project"""
    from database import delete_project
    if not delete_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return {"deleted": True, "id": project_id}


# ======================= RATE LIMITING =======================

# Simple in-memory rate limiter (for production use Redis)
from collections import defaultdict
from time import time

_rate_limit_store = defaultdict(list)
RATE_LIMIT_REQUESTS = 10  # Max requests per minute
RATE_LIMIT_WINDOW = 60  # Window in seconds


def check_rate_limit(client_ip: str, endpoint: str = "default") -> bool:
    """Check if client has exceeded rate limit"""
    key = f"{client_ip}:{endpoint}"
    current_time = time()
    window_start = current_time - RATE_LIMIT_WINDOW
    
    # Clean old entries
    _rate_limit_store[key] = [t for t in _rate_limit_store[key] if t > window_start]
    
    # Check if over limit
    if len(_rate_limit_store[key]) >= RATE_LIMIT_REQUESTS:
        return False
    
    # Record this request
    _rate_limit_store[key].append(current_time)
    return True


@app.get("/api/rate-limit/status")
async def get_rate_limit_status():
    """Get rate limit configuration"""
    return {
        "requests_per_minute": RATE_LIMIT_REQUESTS,
        "window_seconds": RATE_LIMIT_WINDOW,
        "active_clients": len(_rate_limit_store)
    }


# ======================= LBO MODEL GENERATION =======================


@app.post("/api/model/generate-lbo", response_model=ModelResponse)
async def generate_lbo_model_endpoint(request: LBORequest, background_tasks: BackgroundTasks):
    """
    Generate an LBO (Leveraged Buyout) model for a company
    
    This creates a comprehensive LBO model with:
    - Sources & Uses of Funds
    - Debt schedules (Senior, Mezzanine, Subordinated)
    - Operating projections
    - Cash flow and debt paydown
    - Returns analysis (IRR, MoIC, payback period)
    - Sensitivity tables
    """
    job_id = str(uuid.uuid4())
    
    # Initialize job
    jobs[job_id] = {
        "status": "pending",
        "created_at": datetime.now().isoformat(),
        "request": request.dict(),
        "progress": 0,
        "message": "LBO model job queued",
        "model_type": "lbo",
    }
    
    # Start background task
    # Start background task
    if os.environ.get("VERCEL"):
        await _generate_lbo_model_task(job_id, request)
    else:
        background_tasks.add_task(
            _generate_lbo_model_task,
            job_id,
            request,
        )
    
    return ModelResponse(
        job_id=job_id,
        status="pending",
        message="LBO model generation started. Check job status for progress.",
    )


async def _generate_lbo_model_task(job_id: str, request: LBORequest):
    """Background task to generate the LBO model"""
    try:
        jobs[job_id]["status"] = "processing"
        jobs[job_id]["progress"] = 10
        jobs[job_id]["message"] = "Fetching financial data..."
        
        # Step 1: Fetch company data
        logger.info(f"Fetching data for LBO model: {request.symbol}")
        yahoo_data = await fetch_stock_data(request.symbol, request.exchange)
        
        company_info = yahoo_data.get('company_info', {})
        jobs[job_id]["company_name"] = company_info.get('name', request.symbol)
        jobs[job_id]["industry"] = company_info.get('industry', 'Unknown')
        
        jobs[job_id]["progress"] = 30
        jobs[job_id]["message"] = "Preparing LBO assumptions..."
        
        # Step 2: Extract financial data
        income_stmt = yahoo_data.get('income_statement', {})
        key_metrics = yahoo_data.get('key_metrics', {})
        
        # Get base EBITDA and Revenue
        base_ebitda = income_stmt.get('ebitda', key_metrics.get('ebitda', 1000))
        base_revenue = income_stmt.get('revenue', income_stmt.get('total_revenue', 5000))
        
        financial_data = {
            'ebitda': base_ebitda,
            'revenue': base_revenue,
        }
        
        # Step 3: Build LBO assumptions from request
        lbo_assumptions = {
            'entry_multiple': request.entry_multiple,
            'exit_multiple': request.exit_multiple,
            'holding_period': request.holding_period,
            'transaction_fees_pct': request.transaction_fees_pct,
            'financing_fees_pct': request.financing_fees_pct,
            'management_rollover_pct': request.management_rollover_pct,
            'senior_debt_multiple': request.senior_debt_multiple,
            'senior_interest_rate': request.senior_interest_rate,
            'senior_amort_years': request.senior_amort_years,
            'mezz_debt_multiple': request.mezz_debt_multiple,
            'mezz_interest_rate': request.mezz_interest_rate,
            'mezz_pik_rate': request.mezz_pik_rate,
            'sub_debt_multiple': request.sub_debt_multiple,
            'sub_interest_rate': request.sub_interest_rate,
            'revenue_growth': request.revenue_growth,
            'ebitda_margin': request.ebitda_margin,
            'capex_pct': request.capex_pct,
            'nwc_pct': request.nwc_pct,
            'tax_rate': request.tax_rate,
        }
        
        industry_info = {
            'industry': company_info.get('industry', 'General'),
            'sector': company_info.get('sector', 'General'),
        }
        
        jobs[job_id]["progress"] = 50
        jobs[job_id]["message"] = "Generating LBO model structure..."
        
        # Step 4: Generate LBO Excel file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{request.symbol}_LBO_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        
        jobs[job_id]["progress"] = 70
        jobs[job_id]["message"] = "Building Excel model with debt schedules..."
        
        generate_lbo_model(
            company_name=company_info.get('name', request.symbol),
            financial_data=financial_data,
            lbo_assumptions=lbo_assumptions,
            industry_info=industry_info,
            output_path=output_path,
        )
        
        jobs[job_id]["progress"] = 90
        jobs[job_id]["message"] = "Finalizing LBO model..."
        
        # Complete
        jobs[job_id]["status"] = "completed"
        jobs[job_id]["progress"] = 100
        jobs[job_id]["message"] = "LBO model generated successfully!"
        jobs[job_id]["file_path"] = output_path
        jobs[job_id]["filename"] = filename
        jobs[job_id]["download_url"] = f"/api/download/{job_id}"
        jobs[job_id]["lbo_summary"] = {
            "entry_ev": base_ebitda * request.entry_multiple,
            "total_debt": base_ebitda * (request.senior_debt_multiple + request.mezz_debt_multiple + request.sub_debt_multiple),
            "holding_period": request.holding_period,
        }
        
        logger.info(f"LBO model generated successfully: {filename}")
        
    except Exception as e:
        logger.error(f"Error generating LBO model for job {job_id}: {e}")
        jobs[job_id]["status"] = "failed"
        jobs[job_id]["message"] = f"Error: {str(e)}"
        jobs[job_id]["progress"] = 0


# ======================= M&A MODEL GENERATION =======================

@app.post("/api/model/generate-ma", response_model=ModelResponse)
async def generate_ma_model_endpoint(request: MARequest, background_tasks: BackgroundTasks):
    """
    Generate an M&A (Merger & Acquisition) model for two companies
    
    This creates a comprehensive M&A model with:
    - Transaction Summary (Sources & Uses)
    - Standalone Financials
    - Pro Forma Combined Projections
    - Accretion/Dilution Analysis
    - Synergy Modeling
    - Sensitivity Tables
    """
    job_id = str(uuid.uuid4())
    
    # Initialize job
    jobs[job_id] = {
        "status": "pending",
        "created_at": datetime.now().isoformat(),
        "request": request.dict(),
        "progress": 0,
        "message": "M&A model job queued",
        "model_type": "ma",
    }
    
    # Start background task
    # Start background task
    if os.environ.get("VERCEL"):
        await _generate_ma_model_task(job_id, request)
    else:
        background_tasks.add_task(
            _generate_ma_model_task,
            job_id,
            request,
        )
    
    return ModelResponse(
        job_id=job_id,
        status="pending",
        message="M&A model generation started. Check job status for progress.",
    )


async def _generate_ma_model_task(job_id: str, request: MARequest):
    """Background task to generate the M&A model"""
    try:
        jobs[job_id]["status"] = "processing"
        jobs[job_id]["progress"] = 10
        jobs[job_id]["message"] = "Fetching acquirer data..."
        
        # Step 1: Fetch acquirer data
        logger.info(f"Fetching data for M&A model: {request.acquirer_symbol} + {request.target_symbol}")
        acquirer_yahoo = await fetch_stock_data(request.acquirer_symbol, request.exchange)
        
        jobs[job_id]["progress"] = 25
        jobs[job_id]["message"] = "Fetching target data..."
        
        # Step 2: Fetch target data
        target_yahoo = await fetch_stock_data(request.target_symbol, request.exchange)
        
        jobs[job_id]["progress"] = 40
        jobs[job_id]["message"] = "Preparing M&A assumptions..."
        
        # Step 3: Extract financials
        acq_info = acquirer_yahoo.get('company_info', {})
        acq_income = acquirer_yahoo.get('income_statement', {})
        acq_metrics = acquirer_yahoo.get('key_metrics', {})
        
        tgt_info = target_yahoo.get('company_info', {})
        tgt_income = target_yahoo.get('income_statement', {})
        tgt_metrics = target_yahoo.get('key_metrics', {})
        
        acquirer_data = {
            'name': acq_info.get('name', request.acquirer_symbol),
            'revenue': acq_income.get('revenue', acq_income.get('total_revenue', 10000)),
            'ebitda': acq_income.get('ebitda', acq_metrics.get('ebitda', 2000)),
            'net_income': acq_income.get('net_income', 1000),
            'shares_outstanding': acq_metrics.get('shares_outstanding', 500),
            'current_price': acq_info.get('current_price', 200),
        }
        
        target_data = {
            'name': tgt_info.get('name', request.target_symbol),
            'revenue': tgt_income.get('revenue', tgt_income.get('total_revenue', 2000)),
            'ebitda': tgt_income.get('ebitda', tgt_metrics.get('ebitda', 400)),
            'net_income': tgt_income.get('net_income', 200),
            'shares_outstanding': tgt_metrics.get('shares_outstanding', 100),
            'current_price': tgt_info.get('current_price', 100),
            'total_debt': tgt_metrics.get('total_debt', 0),
        }
        
        jobs[job_id]["company_name"] = f"{request.acquirer_symbol} + {request.target_symbol}"
        
        # Step 4: Build transaction assumptions
        transaction_assumptions = {
            'offer_premium': request.offer_premium,
            'percent_stock': request.percent_stock,
            'percent_cash': request.percent_cash,
            'transaction_fees_pct': request.transaction_fees_pct,
            'financing_rate': request.financing_rate,
            'synergies_revenue': request.synergies_revenue,
            'synergies_cost': request.synergies_cost,
            'integration_costs': request.integration_costs,
            'synergy_phase_in_year1': request.synergy_phase_in_year1,
            'synergy_phase_in_year2': request.synergy_phase_in_year2,
            'synergy_phase_in_year3': request.synergy_phase_in_year3,
            'acquirer_growth_rate': request.acquirer_growth_rate,
            'target_growth_rate': request.target_growth_rate,
        }
        
        jobs[job_id]["progress"] = 60
        jobs[job_id]["message"] = "Generating M&A model structure..."
        
        # Step 5: Generate M&A Excel file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{request.acquirer_symbol}_{request.target_symbol}_MA_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        
        jobs[job_id]["progress"] = 75
        jobs[job_id]["message"] = "Building accretion/dilution analysis..."
        
        generate_ma_model(
            acquirer_data=acquirer_data,
            target_data=target_data,
            transaction_assumptions=transaction_assumptions,
            output_path=output_path,
        )
        
        jobs[job_id]["progress"] = 90
        jobs[job_id]["message"] = "Finalizing M&A model..."
        
        # Complete
        jobs[job_id]["status"] = "completed"
        jobs[job_id]["progress"] = 100
        jobs[job_id]["message"] = "M&A model generated successfully!"
        jobs[job_id]["file_path"] = output_path
        jobs[job_id]["filename"] = filename
        jobs[job_id]["download_url"] = f"/api/download/{job_id}"
        jobs[job_id]["ma_summary"] = {
            "acquirer": request.acquirer_symbol,
            "target": request.target_symbol,
            "offer_premium": request.offer_premium,
            "consideration_mix": f"{request.percent_stock*100:.0f}% Stock / {request.percent_cash*100:.0f}% Cash",
        }
        
        logger.info(f"M&A model generated successfully: {filename}")
        
    except Exception as e:
        logger.error(f"Error generating M&A model for job {job_id}: {e}")
        jobs[job_id]["status"] = "failed"
        jobs[job_id]["message"] = f"Error: {str(e)}"
        jobs[job_id]["progress"] = 0


# ======================= STOCK DATABASE ENDPOINTS =======================

@app.get("/api/stocks")
async def list_all_stocks(sector: Optional[str] = None):
    """Get all available stocks, optionally filtered by sector"""
    if sector:
        stocks = get_stocks_by_sector(sector)
    else:
        stocks = get_all_stocks()
    
    return {
        "count": len(stocks),
        "stocks": stocks,
    }


@app.get("/api/stocks/search/{query}")
async def search_for_stocks(query: str):
    """Search stocks by symbol or name"""
    results = search_stocks(query)
    return {
        "count": len(results),
        "results": results,
    }


@app.get("/api/sectors")
async def list_sectors():
    """Get all available sectors"""
    sectors = get_sectors()
    return {
        "sectors": sectors,
        "count": len(sectors),
    }


# ======================= RAW DATA MODEL GENERATION =======================

@app.post("/api/model/generate-raw")
async def generate_model_from_raw_data(request: RawDataRequest, background_tasks: BackgroundTasks):
    """
    Generate a financial model from raw data input
    
    This allows users to provide their own financial data instead of scraping.
    """
    job_id = str(uuid.uuid4())
    
    # Initialize job
    jobs[job_id] = {
        "status": "pending",
        "created_at": datetime.now().isoformat(),
        "request": request.dict(),
        "progress": 0,
        "message": "Job queued",
    }
    
    # Start background task
    # Start background task
    if os.environ.get("VERCEL"):
        await _generate_model_from_raw_task(
            job_id,
            request.company_name,
            request.industry,
            request.forecast_years,
            request.historical_data,
            request.assumptions,
        )
    else:
        background_tasks.add_task(
            _generate_model_from_raw_task,
            job_id,
            request.company_name,
            request.industry,
            request.forecast_years,
            request.historical_data,
            request.assumptions,
        )
    
    return ModelResponse(
        job_id=job_id,
        status="pending",
        message="Model generation from raw data started.",
    )


async def _generate_model_from_raw_task(
    job_id: str,
    company_name: str,
    industry: str,
    forecast_years: int,
    historical_data: Dict[str, Any],
    assumptions: Dict[str, float],
):
    """Background task to generate model from raw data"""
    try:
        jobs[job_id]["status"] = "processing"
        jobs[job_id]["progress"] = 20
        jobs[job_id]["message"] = "Processing raw data..."
        
        # Build financial data structure
        financial_data = {
            'company_info': {
                'name': company_name,
                'sector': industry,
                'industry': industry,
            },
            'income_statement': historical_data.get('income_statement', {}),
            'balance_sheet': historical_data.get('balance_sheet', {}),
            'cash_flow': historical_data.get('cash_flow', {}),
            'key_metrics': historical_data.get('key_metrics', {}),
        }
        
        # Apply user assumptions
        if assumptions:
            financial_data['user_assumptions'] = assumptions
        
        jobs[job_id]["progress"] = 40
        jobs[job_id]["message"] = "Building industry template..."
        
        # Get industry info
        from agents.industry_classifier import INDUSTRY_TEMPLATES
        industry_template = INDUSTRY_TEMPLATES.get(industry, INDUSTRY_TEMPLATES['general'])
        
        industry_info = {
            'industry_name': industry_template['name'],
            'industry_code': industry,
            'model_type': industry_template['model_type'],
            'key_metrics': industry_template['key_metrics'],
        }
        
        jobs[job_id]["company_name"] = company_name
        jobs[job_id]["industry"] = industry_info['industry_name']
        jobs[job_id]["progress"] = 55
        jobs[job_id]["message"] = "Creating model structure..."
        
        # Create model structure
        model_structure = create_model_structure(
            company_name=company_name,
            industry_info=industry_info,
            historical_data=financial_data,
            forecast_years=forecast_years,
        )
        
        # Override with user assumptions
        if assumptions:
            for key_assumption in model_structure.get('key_assumptions', []):
                assumption_name = key_assumption['name'].lower().replace(' ', '_')
                if assumption_name in assumptions:
                    key_assumption['default_value'] = assumptions[assumption_name]
        
        jobs[job_id]["progress"] = 70
        jobs[job_id]["message"] = "Generating Excel model..."
        
        # Generate Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '-', '_')).strip()[:30]
        filename = f"{safe_name}_{industry}_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        
        generate_financial_model(
            company_name=company_name,
            model_structure=model_structure,
            financial_data=financial_data,
            industry_info=industry_info,
            output_path=output_path,
        )
        
        jobs[job_id]["progress"] = 90
        jobs[job_id]["message"] = "Validating model..."
        
        # Validate
        validation_data = {
            'income_statement': financial_data.get('income_statement', {}),
            'balance_sheet': financial_data.get('balance_sheet', {}),
            'cash_flow': financial_data.get('cash_flow', {}),
            'assumptions': assumptions,
        }
        is_valid, errors = validate_financial_model(validation_data, industry)
        
        # Complete
        jobs[job_id]["status"] = "completed"
        jobs[job_id]["progress"] = 100
        jobs[job_id]["message"] = "Model generated successfully from raw data!"
        jobs[job_id]["file_path"] = output_path
        jobs[job_id]["filename"] = filename
        jobs[job_id]["validation"] = {
            "is_valid": is_valid,
            "errors": errors[:5] if errors else [],
        }
        jobs[job_id]["download_url"] = f"/api/download/{job_id}"
        
        logger.info(f"Model generated from raw data: {filename}")
        
    except Exception as e:
        logger.error(f"Error generating model from raw data for job {job_id}: {e}")
        jobs[job_id]["status"] = "failed"
        jobs[job_id]["message"] = f"Error: {str(e)}"
        jobs[job_id]["progress"] = 0


# ======================= EXCEL FILE UPLOAD =======================

@app.post("/api/model/upload-excel")
async def generate_model_from_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    company_name: str = Form("Uploaded Company"),
    industry: str = Form("general"),
    forecast_years: int = Form(5),
):
    """
    Generate a financial model from an uploaded Excel file
    
    The Excel file should have financial data with columns like:
    - Revenue, EBITDA, Net Income (in the income statement sheet or rows)
    - Total Assets, Total Liabilities (in a balance sheet sheet or rows)
    
    The parser will attempt to extract values from common formats.
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    job_id = str(uuid.uuid4())
    
    try:
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        # Parse the Excel file
        extracted_data = _parse_excel_file(tmp_path)
        
        # Clean up temp file
        os.unlink(tmp_path)
        
        # Use extracted company name if available
        if extracted_data.get('company_name'):
            company_name = extracted_data['company_name']
        
        # Initialize job
        jobs[job_id] = {
            "status": "pending",
            "created_at": datetime.now().isoformat(),
            "request": {
                "company_name": company_name,
                "industry": industry,
                "forecast_years": forecast_years,
                "source": "excel_upload",
            },
            "progress": 0,
            "message": "Job queued",
        }
        
        # Start background task
        background_tasks.add_task(
            _generate_model_from_raw_task,
            job_id,
            company_name,
            industry,
            forecast_years,
            extracted_data.get('historical_data', {}),
            extracted_data.get('assumptions', {}),
        )
        
        return ModelResponse(
            job_id=job_id,
            status="pending",
            message="Excel file parsed successfully. Model generation started.",
            company_name=company_name,
        )
        
    except Exception as e:
        logger.error(f"Error processing uploaded Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")


def _parse_excel_file(file_path: str) -> Dict[str, Any]:
    """
    Parse an uploaded Excel file and extract financial data
    
    Supports various common formats:
    - Single sheet with labeled rows (Revenue, EBITDA, etc.)
    - Multiple sheets (Income Statement, Balance Sheet, etc.)
    - Screener.in or similar export formats
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    extracted = {
        'company_name': None,
        'historical_data': {
            'income_statement': {},
            'balance_sheet': {},
        },
        'assumptions': {},
    }
    
    # Common financial data field mappings (case-insensitive)
    income_mappings = {
        'revenue': ['revenue', 'sales', 'total revenue', 'net sales', 'total sales', 'income', 'turnover'],
        'ebitda': ['ebitda', 'operating profit before depreciation', 'pbdit'],
        'net_income': ['net income', 'net profit', 'profit after tax', 'pat', 'bottom line', 'net profit after tax'],
        'operating_income': ['operating income', 'operating profit', 'ebit', 'pbit'],
        'gross_profit': ['gross profit', 'gross margin'],
    }
    
    balance_mappings = {
        'total_assets': ['total assets', 'assets', 'total asset'],
        'total_liabilities': ['total liabilities', 'liabilities', 'total liability', 'total debt'],
        'total_equity': ['total equity', 'shareholders equity', 'shareholder equity', 'net worth', 'networth'],
        'cash': ['cash', 'cash and equivalents', 'cash & equivalents', 'cash and bank'],
        'debt': ['total debt', 'borrowings', 'long term debt', 'short term debt'],
    }
    
    # Try to find company name from first sheet
    first_sheet = wb.active
    for row in first_sheet.iter_rows(min_row=1, max_row=5, max_col=5):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                # Look for company name patterns (usually contains "Ltd" or "Limited" or is in bold/header)
                if any(suffix in val.lower() for suffix in ['ltd', 'limited', 'inc', 'corp', 'company']):
                    extracted['company_name'] = val
                    break
        if extracted['company_name']:
            break
    
    # Search through all sheets for financial data
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lower = sheet_name.lower()
        
        # Determine what type of data this sheet might contain
        is_income_sheet = any(term in sheet_lower for term in ['income', 'profit', 'loss', 'p&l', 'pl', 'revenue'])
        is_balance_sheet = any(term in sheet_lower for term in ['balance', 'asset', 'liability', 'position'])
        
        # Scan the sheet for labeled values
        for row in ws.iter_rows(min_row=1, max_row=100, max_col=10):
            label_cell = row[0] if row else None
            if not label_cell or not label_cell.value:
                continue
            
            label = str(label_cell.value).strip().lower()
            
            # Find the first numeric value in this row
            value = None
            for cell in row[1:]:
                if cell.value is not None:
                    try:
                        value = float(cell.value)
                        break
                    except (ValueError, TypeError):
                        continue
            
            if value is None:
                continue
            
            # Match against income statement fields
            for field, aliases in income_mappings.items():
                if any(alias in label for alias in aliases):
                    extracted['historical_data']['income_statement'][field] = value
                    break
            
            # Match against balance sheet fields
            for field, aliases in balance_mappings.items():
                if any(alias in label for alias in aliases):
                    extracted['historical_data']['balance_sheet'][field] = value
                    break
    
    # Calculate assumptions from extracted data if possible
    income_stmt = extracted['historical_data']['income_statement']
    if income_stmt.get('revenue') and income_stmt.get('ebitda'):
        extracted['assumptions']['ebitda_margin'] = income_stmt['ebitda'] / income_stmt['revenue']
    
    if income_stmt.get('revenue') and income_stmt.get('net_income'):
        extracted['assumptions']['net_margin'] = income_stmt['net_income'] / income_stmt['revenue']
    
    wb.close()
    
    return extracted


@app.get("/api/template/input")
async def get_input_template():
    """Get a sample template for Excel data input"""
    return {
        "description": "Upload an Excel file with financial data",
        "supported_formats": [".xlsx", ".xls"],
        "expected_fields": {
            "income_statement": {
                "revenue": "Annual revenue/sales",
                "ebitda": "Earnings before interest, taxes, depreciation and amortization",
                "net_income": "Net profit after tax",
            },
            "balance_sheet": {
                "total_assets": "Total assets",
                "total_liabilities": "Total liabilities",
            },
        },
        "tips": [
            "Label your rows clearly (e.g., 'Revenue', 'Net Income')",
            "Place values in adjacent columns",
            "Include company name in the first few rows if possible",
            "Multiple sheets are supported (Income Statement, Balance Sheet)",
        ],
    }


# ======================= NEW ENHANCED ENDPOINTS =======================

@app.get("/api/jobs/history")
async def get_job_history(limit: int = 20, offset: int = 0):
    """
    Get job history for re-download
    
    Args:
        limit: Maximum number of jobs to return
        offset: Offset for pagination
    
    Returns:
        List of completed jobs
    """
    try:
        history = db.get_job_history(limit=limit, offset=offset)
        return {
            "jobs": history,
            "count": len(history),
            "limit": limit,
            "offset": offset
        }
    except Exception as e:
        logger.error(f"Error fetching job history: {e}")
        # Fallback to in-memory jobs
        completed = [
            {**v, "id": k} for k, v in jobs.items() 
            if v.get("status") == "completed"
        ]
        return {
            "jobs": completed[:limit],
            "count": len(completed),
            "limit": limit,
            "offset": offset
        }


@app.get("/api/model/preview/{job_id}")
async def get_model_preview(job_id: str):
    """
    Get preview of model metrics before download
    
    Args:
        job_id: Job identifier
    
    Returns:
        Key valuation metrics
    """
    # Try database first
    try:
        metrics = db.get_model_metrics(job_id)
        if metrics:
            return {"job_id": job_id, "metrics": metrics}
    except Exception:
        pass
    
    # Fallback to in-memory
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    return {
        "job_id": job_id,
        "company_name": job.get("company_name"),
        "industry": job.get("industry"),
        "status": job.get("status"),
        "metrics": job.get("metrics", [])
    }


@app.post("/api/ai/smart-assumptions")
async def get_smart_assumptions(
    industry: str = "general",
    company_name: str = "",
    market_cap: float = None
):
    """
    Get AI-powered smart assumptions for financial modeling
    
    Args:
        industry: Industry classification
        company_name: Company name for context
        market_cap: Market cap in crores (optional)
    
    Returns:
        Recommended assumptions with explanations
    """
    try:
        result = await generate_smart_assumptions(
            industry=industry,
            company_name=company_name,
            market_cap=market_cap
        )
        return result
    except Exception as e:
        logger.error(f"Error generating smart assumptions: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ai/commentary")
async def get_ai_commentary(
    company_name: str,
    industry: str,
    valuation_metrics: Dict[str, Any],
    assumptions: Dict[str, Any]
):
    """
    Get AI-generated investment commentary
    
    Args:
        company_name: Company name
        industry: Industry classification
        valuation_metrics: Key metrics (EV, share price, etc.)
        assumptions: Model assumptions
    
    Returns:
        Investment thesis, risks, and recommendation
    """
    try:
        commentary = await generate_valuation_commentary(
            company_name=company_name,
            industry=industry,
            valuation_metrics=valuation_metrics,
            assumptions=assumptions
        )
        return commentary
    except Exception as e:
        logger.error(f"Error generating AI commentary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ai/parse-request")
async def parse_nlp_request(prompt: str):
    """
    Parse natural language request to model parameters
    
    Args:
        prompt: User's natural language request
    
    Returns:
        Extracted parameters for model generation
    """
    try:
        result = await parse_natural_language_request(prompt)
        return result
    except Exception as e:
        logger.error(f"Error parsing NLP request: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/stock/yahoo/{symbol}")
async def get_yahoo_stock_data(symbol: str):
    """
    Get real-time stock data from Yahoo Finance
    
    Args:
        symbol: Stock symbol
    
    Returns:
        Stock information and price history
    """
    try:
        info = get_stock_info(symbol)
        if not info:
            raise HTTPException(status_code=404, detail=f"Stock {symbol} not found")
        
        price_history = get_price_history(symbol, period="1y")
        
        return {
            "info": info,
            "price_history": price_history
        }
    except Exception as e:
        logger.error(f"Error fetching Yahoo data for {symbol}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/stock/historical/{symbol}")
async def get_stock_historical(symbol: str, years: int = 5):
    """
    Get multi-year historical financials
    
    Args:
        symbol: Stock symbol
        years: Number of years
    
    Returns:
        Historical income statement, balance sheet, and cash flow
    """
    try:
        financials = get_historical_financials(symbol, years=years)
        if not financials:
            raise HTTPException(status_code=404, detail=f"No historical data for {symbol}")
        
        return financials
    except Exception as e:
        logger.error(f"Error fetching historical data for {symbol}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/analysis/monte-carlo")
async def run_monte_carlo(
    base_assumptions: Dict[str, float],
    base_valuation: Dict[str, float],
    num_simulations: int = 10000
):
    """
    Run Monte Carlo simulation on valuation
    
    Args:
        base_assumptions: Base case assumptions
        base_valuation: Base case valuation metrics
        num_simulations: Number of simulations (default 10000)
    
    Returns:
        Probability distribution of valuations
    """
    try:
        results = run_monte_carlo_simulation(
            base_assumptions=base_assumptions,
            base_valuation=base_valuation,
            num_simulations=min(num_simulations, 50000)  # Cap at 50k
        )
        return results
    except Exception as e:
        logger.error(f"Error running Monte Carlo: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/{job_id}/{format}")
async def export_model(job_id: str, format: str):
    """
    Export model to different formats
    
    Args:
        job_id: Job identifier
        format: Export format (pdf, pptx, gsheets)
    
    Returns:
        File download or Google Sheets link
    """
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    
    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Model not yet completed")
    
    company_name = job.get("company_name", "Company")
    industry = job.get("industry", "general")
    
    # Get valuation data and assumptions from job
    valuation_data = job.get("valuation_data", {})
    assumptions = job.get("assumptions", {})
    commentary = job.get("commentary")
    
    if format.lower() == "pdf":
        if not pdf_exporter.is_available():
            raise HTTPException(status_code=501, detail="PDF export not available. Install: pip install reportlab")
        
        output_path = os.path.join(OUTPUT_DIR, f"{job_id}_summary.pdf")
        
        success = pdf_exporter.generate_pdf_report(
            output_path=output_path,
            company_name=company_name,
            industry=industry,
            valuation_data=valuation_data,
            assumptions=assumptions,
            commentary=commentary
        )
        
        if success:
            return FileResponse(
                output_path,
                media_type="application/pdf",
                filename=f"{company_name.replace(' ', '_')}_Summary.pdf"
            )
        else:
            raise HTTPException(status_code=500, detail="Failed to generate PDF")
    
    elif format.lower() == "pptx":
        if not pptx_exporter.is_available():
            raise HTTPException(status_code=501, detail="PowerPoint export not available. Install: pip install python-pptx")
        
        output_path = os.path.join(OUTPUT_DIR, f"{job_id}_pitch.pptx")
        
        success = pptx_exporter.generate_pptx_report(
            output_path=output_path,
            company_name=company_name,
            industry=industry,
            valuation_data=valuation_data,
            assumptions=assumptions,
            commentary=commentary
        )
        
        if success:
            return FileResponse(
                output_path,
                media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                filename=f"{company_name.replace(' ', '_')}_Pitch.pptx"
            )
        else:
            raise HTTPException(status_code=500, detail="Failed to generate PowerPoint")
    
    elif format.lower() == "gsheets":
        raise HTTPException(
            status_code=501, 
            detail="Google Sheets export requires additional setup. Configure GOOGLE_APPLICATION_CREDENTIALS."
        )
    
    elif format.lower() == "xlsm":
        # Generate xlsm with VBA modules
        try:
            from exporters.xlsm_generator import create_xlsm_with_vba, get_vba_modules_zip
            
            # Get the actual Excel file path from the job
            xlsx_path = job.get("file_path")
            
            if not xlsx_path or not os.path.exists(xlsx_path):
                raise HTTPException(status_code=404, detail="Original Excel file not found")
            
            xlsm_path = xlsx_path.replace('.xlsx', '_with_vba.xlsm')
            
            success = create_xlsm_with_vba(xlsx_path, xlsm_path)
            
            if success:
                return FileResponse(
                    xlsm_path,
                    media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
                    filename=f"{company_name.replace(' ', '_')}_Model_with_VBA.xlsm"
                )
            else:
                raise HTTPException(status_code=500, detail="Failed to create xlsm file")
        except ImportError:
            raise HTTPException(status_code=501, detail="XLSM generator not available")
    
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {format}")


@app.get("/api/export/{job_id}/vba-modules")
async def get_vba_modules(job_id: str):
    """Download VBA modules as a zip file"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    try:
        from exporters.xlsm_generator import get_vba_modules_zip
        
        company_name = jobs[job_id].get("company_name", "Model")
        zip_path = os.path.join(OUTPUT_DIR, f"{job_id}_vba_modules.zip")
        
        get_vba_modules_zip(zip_path)
        
        return FileResponse(
            zip_path,
            media_type="application/zip",
            filename=f"{company_name.replace(' ', '_')}_VBA_Modules.zip"
        )
    except Exception as e:
        logger.error(f"Failed to create VBA modules zip: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/formats")
async def get_export_formats():
    """Get available export formats"""
    return {
        "formats": [
            {
                "id": "xlsx",
                "name": "Excel",
                "extension": ".xlsx",
                "available": True,
                "description": "Full financial model workbook"
            },
            {
                "id": "xlsm",
                "name": "Excel with VBA",
                "extension": ".xlsm",
                "available": True,
                "description": "Macro-enabled workbook with VBA automation"
            },
            {
                "id": "pdf",
                "name": "PDF Summary",
                "extension": ".pdf",
                "available": pdf_exporter.is_available(),
                "description": "Executive summary report"
            },
            {
                "id": "pptx",
                "name": "PowerPoint",
                "extension": ".pptx",
                "available": pptx_exporter.is_available(),
                "description": "Investor pitch deck"
            },
            {
                "id": "gsheets",
                "name": "Google Sheets",
                "extension": None,
                "available": False,
                "description": "Export to Google Sheets (requires setup)"
            }
        ]
    }


@app.get("/api/preferences/{key}")
async def get_preference(key: str):
    """Get user preference"""
    try:
        value = db.get_preference(key)
        return {"key": key, "value": value}
    except Exception as e:
        return {"key": key, "value": None}


@app.post("/api/preferences/{key}")
async def set_preference(key: str, value: str):
    """Set user preference"""
    try:
        db.set_preference(key, value)
        return {"key": key, "value": value, "success": True}
    except Exception as e:
        logger.error(f"Error setting preference: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# ENHANCEMENT ENDPOINTS
# ============================================================

@app.get("/api/analysis/sensitivity/{job_id}")
async def get_sensitivity_analysis(job_id: str, variation: float = 0.10):
    """
    Get sensitivity analysis for tornado chart visualization
    
    Args:
        job_id: Job identifier
        variation: Percentage variation for each assumption (default 10%)
    """
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Model not ready")
    
    try:
        from analysis.tornado_analysis import calculate_sensitivity
        
        valuation_data = job.get("valuation_data", {})
        assumptions = job.get("assumptions", {})
        
        sensitivity = calculate_sensitivity(valuation_data, assumptions, variation)
        
        return {
            "job_id": job_id,
            "company_name": job.get("company_name"),
            "sensitivity": sensitivity
        }
    except Exception as e:
        logger.error(f"Sensitivity analysis error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/chat/{job_id}")
async def chat_with_model(job_id: str, request: dict):
    """
    AI chat about the financial model
    
    Args:
        job_id: Job identifier
        request: {"message": "user question", "history": []}
    """
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    message = request.get("message", "")
    history = request.get("history", [])
    
    if not message:
        raise HTTPException(status_code=400, detail="Message required")
    
    try:
        from agents.chat_assistant import process_chat_message, get_suggested_questions
        
        response = process_chat_message(message, job, history)
        suggestions = get_suggested_questions(job)
        
        return {
            "job_id": job_id,
            "response": response.get("response", ""),
            "success": response.get("success", False),
            "suggestions": suggestions[:3]
        }
    except Exception as e:
        logger.error(f"Chat error: {e}")
        return {
            "job_id": job_id,
            "response": f"Sorry, I encountered an error: {str(e)}",
            "success": False,
            "suggestions": []
        }


@app.get("/api/chat/{job_id}/suggestions")
async def get_chat_suggestions(job_id: str):
    """Get suggested questions for the model"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    try:
        from agents.chat_assistant import get_suggested_questions
        suggestions = get_suggested_questions(jobs[job_id])
        return {"suggestions": suggestions}
    except Exception as e:
        return {"suggestions": ["What are the key value drivers?", "What are the main risks?"]}


@app.get("/api/analysis/football-field/{job_id}")
async def get_football_field(job_id: str):
    """
    Get football field valuation summary
    
    Aggregates DCF, Monte Carlo, and Comps into visual range
    """
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Model not ready")
    
    try:
        from analysis.football_field import create_football_field
        
        valuation_data = job.get("valuation_data", {})
        monte_carlo = job.get("monte_carlo_results")
        
        football_field = create_football_field(valuation_data, monte_carlo)
        
        return {
            "job_id": job_id,
            "company_name": job.get("company_name"),
            "football_field": football_field
        }
    except Exception as e:
        logger.error(f"Football field error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/templates")
async def get_model_templates():
    """Get available model templates"""
    try:
        import json
        templates_path = os.path.join(os.path.dirname(__file__), "config", "templates.json")
        
        # Try relative path first
        if not os.path.exists(templates_path):
            templates_path = os.path.join(os.path.dirname(__file__), "..", "config", "templates.json")
        
        if os.path.exists(templates_path):
            with open(templates_path, 'r') as f:
                data = json.load(f)
            return data
        else:
            return {"templates": {}, "scenarios": {}}
    except Exception as e:
        logger.error(f"Templates error: {e}")
        return {"templates": {}, "scenarios": {}}


@app.get("/api/templates/{template_id}")
async def get_template(template_id: str):
    """Get specific template by ID"""
    templates = await get_model_templates()
    template = templates.get("templates", {}).get(template_id)
    
    if not template:
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found")
    
    return template


@app.get("/api/stocks/us")
async def get_us_stocks(search: str = None, sector: str = None, limit: int = 50):
    """
    Get US stocks (NYSE/NASDAQ)
    
    Args:
        search: Search term for symbol or name
        sector: Filter by sector
        limit: Max results
    """
    try:
        import json
        us_stocks_path = os.path.join(os.path.dirname(__file__), "data", "us_stocks.json")
        
        if os.path.exists(us_stocks_path):
            with open(us_stocks_path, 'r') as f:
                data = json.load(f)
            
            stocks = data.get("stocks", [])
            
            # Apply filters
            if search:
                search = search.upper()
                stocks = [s for s in stocks if search in s["symbol"] or search in s["name"].upper()]
            
            if sector:
                stocks = [s for s in stocks if sector.lower() in s["sector"].lower()]
            
            return {"stocks": stocks[:limit], "total": len(stocks)}
        else:
            return {"stocks": [], "total": 0}
    except Exception as e:
        logger.error(f"US stocks error: {e}")
        return {"stocks": [], "total": 0}


@app.post("/api/scenarios/{job_id}/save")
async def save_scenario(job_id: str, request: dict):
    """Save current assumptions as a scenario"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    scenario_name = request.get("name", "Saved Scenario")
    
    job = jobs[job_id]
    scenario = {
        "name": scenario_name,
        "assumptions": job.get("assumptions", {}),
        "valuation_data": job.get("valuation_data", {}),
        "created_at": datetime.now().isoformat()
    }
    
    # Store in job
    if "scenarios" not in job:
        job["scenarios"] = []
    job["scenarios"].append(scenario)
    
    return {"success": True, "scenario": scenario}


@app.get("/api/scenarios/{job_id}")
async def get_scenarios(job_id: str):
    """Get saved scenarios for a job"""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    return {"scenarios": jobs[job_id].get("scenarios", [])}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)


